import requests
from openpyxl import load_workbook
import csv
import os
import json
from src import function as func
from src import bo

def process(url,name,path):
    req = requests.get(url)
    url_content = req.content
    tmp_file = './'+name+'.xlsx'
    f = open(tmp_file, 'wb')
    f.write(url_content)
    f.close()

    wb = load_workbook(tmp_file)
    dataset = []
    for sheet in wb:
        for row in sheet.iter_rows():
            data = []
            for cell in row:
                data.append(cell.value)
            dataset.append(data)
    os.remove(tmp_file) 
    output = path+'/'+name+'.csv'
    func.writetofile(output,dataset)


def Extract(path):
    print("====firestation====")
    process("https://data.moi.gov.tw/MoiOD/System/DownloadFile.aspx?DATA=F412EAC2-1953-40BA-8AF3-A7A45EE253E3",'firestation1',path) 
    process("https://data.moi.gov.tw/MoiOD/System/DownloadFile.aspx?DATA=C38B7AC2-E7F3-4DD5-A3F3-88E623B55924",'firestation2',path) 
    


def Transform1(file):
    raw = func.readcsv(file)
    raw = raw[1:]
    data = [] 
    idx = 1
    for r in raw:
        data.append(['1-'+str(idx)]+[r[0],r[1],r[2],r[4],r[5],func.towkt(r[4],r[5])])
        idx = idx +1
        
    return data

def Transform2(file):
    raw = func.readcsv(file)
    raw = raw[1:]
    data = [] 
    idx = 1
    for r in raw:
        data.append(['2-'+str(idx)]+[r[0],r[1],r[2],r[3],r[4],func.towkt(r[3],r[4])])
        idx = idx + 1
    return data


if __name__ == '__main__':
    # Extract
    Extract('./data')
    
    # Transform
    data1 = Transform1('./data/firestation1.csv')
    data2 = Transform2('./data/firestation2.csv')

    # Load
    db = bo.conn("127.0.0.1",13303,"house")
    with open("schema.json") as f:
        schema = json.load(f)
    table = "firestation"
    bo.load(db,table,schema[table],data1)   
    bo.load(db,table,schema[table],data2,delete=False)        


