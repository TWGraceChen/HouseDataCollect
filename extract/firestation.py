import requests
from openpyxl import load_workbook
import csv
import os
import function as func


def process(url,name,path):
    req = requests.get(url)
    url_content = req.content
    tmp_file = './'+name+'.xlsx'
    f = open(tmp_file, 'wb')
    f.write(url_content)
    f.close()

    wb = load_workbook(tmp_file)
    dataset = []
    for sheet in wb:
        for row in sheet.iter_rows():
            data = []
            for cell in row:
                data.append(cell.value)
            dataset.append(data)
    os.remove(tmp_file) 
    output = path+'/'+name+'.csv'
    func.writetofile(output,dataset)


def Extract(path):
    print("====firestation====")
    process("https://data.moi.gov.tw/MoiOD/System/DownloadFile.aspx?DATA=F412EAC2-1953-40BA-8AF3-A7A45EE253E3",'firestation1',path) 
    process("https://data.moi.gov.tw/MoiOD/System/DownloadFile.aspx?DATA=C38B7AC2-E7F3-4DD5-A3F3-88E623B55924",'firestation2',path) 
    

if __name__ == '__main__':
    Extract('../data')
    


